"""
app.py — Dienstplan Web-App (Flask)

Aufruf:
    cd dienstplan
    python app.py
    → http://localhost:5050          (Ansicht für alle)
    → http://<LAN-IP>:5050          (andere Geräte im Netzwerk)

PIN ändern: Umgebungsvariable EDIT_PIN setzen, oder direkt hier:
    EDIT_PIN = "1234"
"""

import base64
import hashlib
import json
import os
import secrets
import socket
import sys
import threading
import urllib.request
from datetime import date
from functools import wraps
from pathlib import Path

from flask import Flask, jsonify, request, render_template, session, Response
from flask_cors import CORS

DIENSTPLAN_DIR = Path(__file__).resolve().parent
REPO_DIR       = DIENSTPLAN_DIR.parent
sys.path.insert(0, str(DIENSTPLAN_DIR))

from scheduler import generate_week
from vacations import load_vacations
from config    import ALL_PERSONS, ARZTE, TFAS, DAYS

# Excel-Datei: erst im selben Verzeichnis suchen (Deployment), dann im Repo-Root (lokal)
_urlaub_candidates = [
    DIENSTPLAN_DIR / "Urlaubsplanung 2026.xlsx",
    REPO_DIR / "Urlaubsplanung 2026.xlsx",
]
URLAUB_PATH = next((p for p in _urlaub_candidates if p.exists()), _urlaub_candidates[-1])
if not URLAUB_PATH.exists():
    print(f"WARNUNG: Urlaubsplanung nicht gefunden. Gesucht: {_urlaub_candidates}")

OVERRIDES_PATH     = DIENSTPLAN_DIR / "overrides.json"
VAC_OVERRIDES_PATH = DIENSTPLAN_DIR / "vacation_overrides.json"
OT_ROTATION_PATH   = DIENSTPLAN_DIR / "overtime_rotation.json"

# PIN für den Bearbeitungsmodus (Umgebungsvariable EDIT_PIN, Fallback "1234")
EDIT_PIN = os.environ.get("EDIT_PIN", "1234")

# ---------------------------------------------------------------------------
# GitHub-Persistenz (verhindert Datenverlust bei Server-Neustart auf Render)
# Env-Vars: GITHUB_TOKEN, GITHUB_REPO (optional, Default: soeren1508/dienstplan)
# ---------------------------------------------------------------------------
GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN", "")
GITHUB_REPO  = os.environ.get("GITHUB_REPO", "soeren1508/dienstplan")
GITHUB_API   = "https://api.github.com"

def _gh_headers() -> dict:
    return {
        "Authorization": f"token {GITHUB_TOKEN}",
        "Accept": "application/vnd.github.v3+json",
        "User-Agent": "DienstplanApp/1.0",
        "Content-Type": "application/json",
    }

def _github_pull(filename: str) -> str | None:
    """Liest eine Datei aus dem GitHub-Repo. Gibt den Inhalt oder None zurück."""
    if not GITHUB_TOKEN:
        return None
    url = f"{GITHUB_API}/repos/{GITHUB_REPO}/contents/{filename}"
    try:
        req = urllib.request.Request(url, headers=_gh_headers())
        with urllib.request.urlopen(req, timeout=10) as r:
            data = json.loads(r.read())
            return base64.b64decode(data["content"]).decode("utf-8")
    except Exception as e:
        print(f"[GitHub] Pull {filename} fehlgeschlagen: {e}")
        return None

def _github_push(filename: str, content: str, message: str = "Auto-save"):
    """Schreibt eine Datei in das GitHub-Repo (im Hintergrund)."""
    if not GITHUB_TOKEN:
        return
    def _do():
        url = f"{GITHUB_API}/repos/{GITHUB_REPO}/contents/{filename}"
        headers = _gh_headers()
        # Aktuellen SHA holen (nötig für Update)
        sha = None
        try:
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=10) as r:
                sha = json.loads(r.read()).get("sha")
        except Exception:
            pass
        body: dict = {
            "message": message,
            "content": base64.b64encode(content.encode("utf-8")).decode(),
            "branch": "main",
        }
        if sha:
            body["sha"] = sha
        try:
            req = urllib.request.Request(
                url, data=json.dumps(body).encode(), headers=headers, method="PUT"
            )
            with urllib.request.urlopen(req, timeout=15):
                pass
            print(f"[GitHub] {filename} gespeichert ✓")
        except Exception as e:
            print(f"[GitHub] Push {filename} fehlgeschlagen: {e}")
    threading.Thread(target=_do, daemon=True).start()

def _sync_from_github():
    """Beim Start: Overrides von GitHub laden (aktuellster Stand)."""
    for path, filename in [
        (OVERRIDES_PATH,     "overrides.json"),
        (VAC_OVERRIDES_PATH, "vacation_overrides.json"),
        (OT_ROTATION_PATH,   "overtime_rotation.json"),
    ]:
        content = _github_pull(filename)
        if content:
            try:
                json.loads(content)   # Validierung
                path.write_text(content, encoding="utf-8")
                print(f"[GitHub] {filename} geladen ✓")
            except Exception as e:
                print(f"[GitHub] {filename} ungültig: {e}")

app = Flask(__name__)
CORS(app, supports_credentials=True)

# SECRET_KEY: deterministisch ableiten, damit Sessions auch nach Server-Neustarts
# (Render Auto-Deploy nach GitHub-Push!) gültig bleiben.
# Priorität: Env-Var SECRET_KEY → Hash aus PIN → zufällig (nur Notfall)
_key_base   = os.environ.get("SECRET_KEY") or ("dienstplan-2026-" + EDIT_PIN)
app.secret_key = hashlib.sha256(_key_base.encode()).hexdigest()

from datetime import timedelta
app.config.update(
    PERMANENT_SESSION_LIFETIME = timedelta(hours=8),
    SESSION_COOKIE_SAMESITE    = "Lax",
    SESSION_COOKIE_HTTPONLY    = True,
)

# _sync_from_github() und _generate_all() werden NACH deren Definition aufgerufen
# (siehe unten, nach _generate_all)


# ---------------------------------------------------------------------------
# Auth-Hilfsfunktionen
# ---------------------------------------------------------------------------

def _pin_hash(pin: str) -> str:
    return hashlib.sha256(pin.encode()).hexdigest()

def _is_authenticated() -> bool:
    return session.get("authenticated") is True

def require_auth(f):
    """Decorator: gibt 403 zurück wenn nicht eingeloggt."""
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not _is_authenticated():
            return jsonify({"ok": False, "error": "Nicht autorisiert"}), 403
        return f(*args, **kwargs)
    return wrapper


# ---------------------------------------------------------------------------
# Plan-Cache
# ---------------------------------------------------------------------------
_plan_cache: dict[int, dict] = {}


# ---------------------------------------------------------------------------
# Urlaubs-Overrides (manuelle Ergänzungen / Streichungen über die App)
# ---------------------------------------------------------------------------

def _load_vac_overrides() -> dict:
    if VAC_OVERRIDES_PATH.exists():
        with open(VAC_OVERRIDES_PATH, encoding="utf-8") as f:
            return json.load(f)
    return {}


def _save_vac_overrides(data: dict):
    content = json.dumps(data, indent=2, ensure_ascii=False)
    VAC_OVERRIDES_PATH.write_text(content, encoding="utf-8")
    _github_push("vacation_overrides.json", content, "Auto-save: Urlaubsänderung")


def _get_all_vac():
    """Excel-Urlaub + manuelle Korrekturen zusammenführen."""
    all_vac = load_vacations(URLAUB_PATH)
    for person, changes in _load_vac_overrides().items():
        if person not in all_vac:
            all_vac[person] = set()
        for iso in changes.get("add", []):
            all_vac[person].add(date.fromisoformat(iso))
        for iso in changes.get("remove", []):
            all_vac[person].discard(date.fromisoformat(iso))
    return all_vac


def _generate_all():
    global _plan_cache
    all_vac = _get_all_vac()
    auffl   = {p: 0 for p in ALL_PERSONS}
    plans   = {}

    # KW17–23: aus finalisertem Excel laden (unveränderter Referenzplan)
    # Die Datei hat je ein Blatt "KW17"…"KW23"; Person in Spalte A, Mo–Sa in B–G.
    _excel_kw17_23 = DIENSTPLAN_DIR / "Dienstplan_KW17-23.xlsx"
    _SKIP_ROWS = {"ÄRZTE", "TFAS", "Person"}   # Abschnitt-Header überspringen

    def _load_kw_sheet(wb, kw: int) -> dict | None:
        """Liest ein KW-Blatt und gibt {person: [mo..sa]} zurück, oder None."""
        sheet_name = f"KW{kw}"
        if sheet_name not in wb.sheetnames:
            return None
        ws = wb[sheet_name]
        result = {}
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=7, values_only=True):
            cell0 = str(row[0]).strip() if row[0] is not None else ""
            if not cell0 or cell0 in _SKIP_ROWS or cell0.startswith("Dienstplan") or cell0.startswith("Farben"):
                continue
            if cell0 in ALL_PERSONS:
                result[cell0] = [str(v).strip() if v is not None else "–" for v in row[1:7]]
        return result if result else None

    if _excel_kw17_23.exists():
        try:
            import openpyxl as _opx
            _wb = _opx.load_workbook(_excel_kw17_23, data_only=True)
            for kw in range(17, 24):
                raw = _load_kw_sheet(_wb, kw)
                if raw:
                    plans[kw] = {p: raw.get(p, ["–"] * 6) for p in ALL_PERSONS}
                else:
                    print(f"[WARNUNG] KW{kw} nicht in Excel gefunden – generiere")
                    plans[kw] = generate_week(kw, all_vac, auffl)
        except Exception as e:
            print(f"[WARNUNG] Dienstplan_KW17-23.xlsx konnte nicht geladen werden: {e}")
            for kw in range(17, 24):
                plans[kw] = generate_week(kw, all_vac, auffl)
    else:
        print("[WARNUNG] Dienstplan_KW17-23.xlsx nicht gefunden – generiere KW17–23")
        for kw in range(17, 24):
            plans[kw] = generate_week(kw, all_vac, auffl)

    # KW24–52: auto-generiert
    for kw in range(24, 53):
        plans[kw] = generate_week(kw, all_vac, auffl)

    _plan_cache = plans


# Beim Start (auch unter Gunicorn): Overrides von GitHub laden, Plan generieren
_sync_from_github()
_generate_all()


def _iso_kw(d) -> int:
    """ISO-Kalenderwoche für ein date-Objekt."""
    return d.isocalendar()[1]


def _load_overrides() -> dict:
    if OVERRIDES_PATH.exists():
        with open(OVERRIDES_PATH, encoding="utf-8") as f:
            return json.load(f)
    return {}


def _save_overrides(data: dict):
    content = json.dumps(data, indent=2, ensure_ascii=False)
    OVERRIDES_PATH.write_text(content, encoding="utf-8")
    _github_push("overrides.json", content, "Auto-save: Dienstplan-Änderung")


def _plan_for_kw(kw: int, overrides: dict) -> dict:
    base  = _plan_cache.get(kw, {})
    kw_ov = overrides.get(str(kw), {})
    result = {}
    for person in ALL_PERSONS:
        row = list(base.get(person, ["–"] * 6))
        for di_str, val in kw_ov.get(person, {}).items():
            row[int(di_str)] = val
        result[person] = row
    return result


# ---------------------------------------------------------------------------
# Routen — Öffentlich (View)
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/manifest.json")
def manifest():
    data = {
        "name": "Dienstplan 2026",
        "short_name": "Dienstplan",
        "description": "Tierarztpraxis Hamburg – Dienstplan 2026",
        "start_url": "/",
        "display": "standalone",
        "background_color": "#f5f6fa",
        "theme_color": "#1e293b",
        "orientation": "portrait-primary",
        "icons": [
            {"src": "/static/logo.png", "sizes": "any", "type": "image/png", "purpose": "any maskable"},
        ],
        "categories": ["productivity", "medical"],
    }
    return Response(
        __import__("json").dumps(data, ensure_ascii=False),
        mimetype="application/manifest+json"
    )


@app.route("/api/me")
def api_me():
    """Gibt Auth-Status zurück (wird beim Seitenlade abgefragt)."""
    return jsonify({"authenticated": _is_authenticated()})


@app.route("/api/plan/<int:kw>")
def api_plan(kw):
    if not _plan_cache:
        _generate_all()
    overrides    = _load_overrides()
    plan         = _plan_for_kw(kw, overrides)
    kw_ov        = overrides.get(str(kw), {})
    from scheduler import week_dates
    dates        = week_dates(kw)
    date_labels  = [d.strftime("%-d.%-m.") for d in dates]
    return jsonify({
        "kw":        kw,
        "dates":     date_labels,
        "plan":      plan,
        "days":      DAYS,
        "persons":   ALL_PERSONS,
        "arzte":     ARZTE,
        "tfas":      TFAS,
        "overrides": kw_ov,
    })


@app.route("/api/validate/<int:kw>")
def api_validate(kw):
    if not _plan_cache:
        _generate_all()
    overrides = _load_overrides()
    plan      = _plan_for_kw(kw, overrides)
    issues    = _validate_plan(plan, kw)
    return jsonify(issues)


@app.route("/api/validate/<int:kw>/suggestions")
def api_suggestions(kw):
    if not _plan_cache:
        _generate_all()
    overrides   = _load_overrides()
    plan        = _plan_for_kw(kw, overrides)
    issues      = _validate_plan(plan, kw)
    suggestions = _suggest_fixes(plan, issues, kw)
    return jsonify(suggestions)


# ---------------------------------------------------------------------------
# Routen — Auth
# ---------------------------------------------------------------------------

@app.route("/api/auth", methods=["POST"])
def api_auth():
    data = request.get_json() or {}
    pin  = str(data.get("pin", ""))
    if pin == EDIT_PIN:
        session["authenticated"] = True
        session.permanent = True    # 8 h (PERMANENT_SESSION_LIFETIME)
        return jsonify({"ok": True})
    return jsonify({"ok": False, "error": "Falscher PIN"}), 401


@app.route("/api/auth", methods=["DELETE"])
def api_logout():
    session.clear()
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Routen — Geschützt (Edit)
# ---------------------------------------------------------------------------

@app.route("/api/override", methods=["POST"])
@require_auth
def api_override():
    data   = request.get_json()
    kw     = str(data["kw"])
    person = data["person"]
    di     = str(data["di"])
    value  = data["value"]

    overrides = _load_overrides()
    if value == "__reset__":
        overrides.get(kw, {}).get(person, {}).pop(di, None)
        if not overrides.get(kw, {}).get(person):
            overrides.get(kw, {}).pop(person, None)
        if not overrides.get(kw):
            overrides.pop(kw, None)
    else:
        overrides.setdefault(kw, {}).setdefault(person, {})[di] = value

    _save_overrides(overrides)
    return jsonify({"ok": True})


@app.route("/api/regenerate", methods=["POST"])
@require_auth
def api_regenerate():
    _generate_all()
    return jsonify({"ok": True})


@app.route("/api/overrides/clear/<int:kw>", methods=["POST"])
@require_auth
def api_clear_kw(kw):
    overrides = _load_overrides()
    overrides.pop(str(kw), None)
    _save_overrides(overrides)
    return jsonify({"ok": True})


@app.route("/api/absence", methods=["POST"])
@require_auth
def api_absence():
    """
    Trägt eine Abwesenheit (Krank / Frei / Urlaub) als Override ein und
    gibt sofort die Validierungs-Issues für den betroffenen Tag zurück.

    Body: { kw, person, di, grund }
    Antwort: { ok, issues_day: [...], issues_cell: {...} }
    """
    if not _plan_cache:
        _generate_all()
    data   = request.get_json()
    kw     = int(data["kw"])
    person = data["person"]
    di     = int(data["di"])
    grund  = data.get("grund", "Krank")   # "Krank", "Urlaub", "Frei"

    if person not in ALL_PERSONS:
        return jsonify({"ok": False, "error": f"Unbekannte Person: {person}"}), 400
    if di < 0 or di > 5:
        return jsonify({"ok": False, "error": "Ungültiger Tag"}), 400

    overrides = _load_overrides()
    overrides.setdefault(str(kw), {}).setdefault(person, {})[str(di)] = grund
    _save_overrides(overrides)

    plan        = _plan_for_kw(kw, overrides)
    issues      = _validate_plan(plan, kw)
    suggestions = _suggest_fixes(plan, issues, kw)

    return jsonify({
        "ok":          True,
        "issues_day":  issues["day"].get(str(di), []),
        "issues_cell": {
            p: issues["cell"][p].get(str(di), [])
            for p in ALL_PERSONS
            if issues["cell"][p].get(str(di))
        },
        "suggestions":   suggestions,
        "total_issues":  sum(len(v) for v in issues["day"].values()),
    })


@app.route("/api/swap", methods=["POST"])
@require_auth
def api_swap():
    """
    Tauscht die Dienste von zwei Personen an einem Tag.
    Body: { kw, di, person_a, person_b }
    """
    if not _plan_cache:
        _generate_all()
    data     = request.get_json()
    kw       = int(data["kw"])
    di       = int(data["di"])
    person_a = data["person_a"]
    person_b = data["person_b"]

    if person_a not in ALL_PERSONS or person_b not in ALL_PERSONS:
        return jsonify({"ok": False, "error": "Unbekannte Person"}), 400
    if person_a == person_b:
        return jsonify({"ok": False, "error": "Gleiche Person gewählt"}), 400

    overrides = _load_overrides()
    plan      = _plan_for_kw(kw, overrides)

    val_a = str(plan[person_a][di]).replace(" + Auffüllen", "").strip()
    val_b = str(plan[person_b][di]).replace(" + Auffüllen", "").strip()

    # Overrides setzen
    overrides.setdefault(str(kw), {}).setdefault(person_a, {})[str(di)] = val_b
    overrides.setdefault(str(kw), {}).setdefault(person_b, {})[str(di)] = val_a
    _save_overrides(overrides)

    plan_new    = _plan_for_kw(kw, overrides)
    issues      = _validate_plan(plan_new, kw)
    suggestions = _suggest_fixes(plan_new, issues, kw)

    return jsonify({
        "ok":   True,
        "swap": {
            "person_a": person_a, "before_a": val_a, "after_a": val_b,
            "person_b": person_b, "before_b": val_b, "after_b": val_a,
        },
        "issues":       issues,
        "suggestions":  suggestions,
        "total_issues": sum(len(v) for v in issues["day"].values()),
    })


@app.route("/api/vacation-day", methods=["POST"])
@require_auth
def api_vacation_day():
    """
    Fügt einen Urlaubstag hinzu oder entfernt ihn — triggert komplette Neuplanung.
    Body: { person, date: "2026-05-11", action: "add"|"remove" }
    """
    data    = request.get_json()
    person  = data["person"]
    iso     = data["date"]
    action  = data["action"]   # "add" | "remove"

    if person not in ALL_PERSONS:
        return jsonify({"ok": False, "error": "Unbekannte Person"}), 400

    vac_ov     = _load_vac_overrides()
    person_ov  = vac_ov.setdefault(person, {"add": [], "remove": []})

    if action == "add":
        if iso not in person_ov["add"]:
            person_ov["add"].append(iso)
        if iso in person_ov.get("remove", []):
            person_ov["remove"].remove(iso)
    elif action == "remove":
        if iso not in person_ov.get("remove", []):
            person_ov.setdefault("remove", []).append(iso)
        if iso in person_ov.get("add", []):
            person_ov["add"].remove(iso)
    else:
        return jsonify({"ok": False, "error": "action muss 'add' oder 'remove' sein"}), 400

    _save_vac_overrides(vac_ov)
    _generate_all()   # Neuplanung mit aktualisierten Urlaubs-Daten

    kw        = _iso_kw(date.fromisoformat(iso))
    overrides = _load_overrides()
    plan      = _plan_for_kw(kw, overrides)
    issues    = _validate_plan(plan, kw)
    suggestions = _suggest_fixes(plan, issues, kw)

    from scheduler import week_dates
    dates      = week_dates(kw)
    date_labels = [d.strftime("%-d.%-m.") for d in dates]

    return jsonify({
        "ok":           True,
        "kw":           kw,
        "dates":        date_labels,
        "plan":         plan,
        "issues":       issues,
        "suggestions":  suggestions,
        "total_issues": sum(len(v) for v in issues["day"].values()),
    })


@app.route("/api/apply-suggestions", methods=["POST"])
@require_auth
def api_apply_suggestions():
    """
    Wendet einen oder mehrere Vorschläge als Overrides an.
    Body: { kw, overrides: [{person, di, value}, ...] }
    """
    if not _plan_cache:
        _generate_all()
    data        = request.get_json()
    kw          = int(data["kw"])
    new_overrides_list = data["overrides"]

    overrides = _load_overrides()
    for ov in new_overrides_list:
        overrides.setdefault(str(kw), {}).setdefault(ov["person"], {})[str(ov["di"])] = ov["value"]
    _save_overrides(overrides)

    plan        = _plan_for_kw(kw, overrides)
    issues      = _validate_plan(plan, kw)
    suggestions = _suggest_fixes(plan, issues, kw)

    return jsonify({
        "ok":           True,
        "issues":       issues,
        "suggestions":  suggestions,
        "total_issues": sum(len(v) for v in issues["day"].values()),
    })


# ---------------------------------------------------------------------------
# Lösungsvorschläge für Regelverstoße
# ---------------------------------------------------------------------------

def _suggest_fixes(plan: dict, issues: dict, kw: int) -> list:
    """
    Analysiert Regel-Verstöße und gibt konkrete Override-Vorschläge zurück.
    Format je Vorschlag:
      { day: int, msg: str, description: str,
        overrides: [{person, di, value}, ...] }
    """
    def _is_free(val):
        s = str(val) if val else "–"
        return s in ("–", "Urlaub", "Feiertag", "SCHULE", "FREI_SCHUTZ",
                     "Abschlussprüfung", "Krank", "Frei") or "Frei" in s

    suggestions = []

    for di_str, msgs in issues["day"].items():
        di  = int(di_str)
        row = {p: str(plan[p][di]) for p in ALL_PERSONS}

        def tfa_in_shift(shift):
            return [p for p in TFAS if shift in row[p] and not _is_free(row[p])]

        for msg in msgs:
            # ── Anmeldung fehlt ──────────────────────────────────────────────
            if "FD: Anmeldung fehlt" in msg:
                cands = [p for p in tfa_in_shift("FD") if "Anmeldung" not in row[p]]
                for c in cands[:3]:
                    old = row[c].replace(" + Auffüllen", "")
                    suggestions.append({
                        "day": di, "msg": msg,
                        "description": f"{DAYS[di]}: {c}  {old} → FD: Anmeldung",
                        "overrides": [{"person": c, "di": di, "value": "FD: Anmeldung"}],
                    })

            elif "SD: Anmeldung fehlt" in msg:
                cands = [p for p in tfa_in_shift("SD") if "Anmeldung" not in row[p]]
                for c in cands[:3]:
                    old = row[c].replace(" + Auffüllen", "")
                    suggestions.append({
                        "day": di, "msg": msg,
                        "description": f"{DAYS[di]}: {c}  {old} → SD: Anmeldung",
                        "overrides": [{"person": c, "di": di, "value": "SD: Anmeldung"}],
                    })

            # ── Ulf-Assistenz fehlt ──────────────────────────────────────────
            elif "FD: Assistenz" in msg and "Ulf" in msg:
                cands = [p for p in tfa_in_shift("FD") if "Assistenz" not in row[p]]
                for c in cands[:3]:
                    old = row[c].replace(" + Auffüllen", "")
                    suggestions.append({
                        "day": di, "msg": msg,
                        "description": f"{DAYS[di]}: {c}  {old} → FD: Assistenz Ulf",
                        "overrides": [{"person": c, "di": di, "value": "FD: Assistenz Ulf"}],
                    })

            elif "SD: Assistenz" in msg and "Ulf" in msg:
                cands = [p for p in tfa_in_shift("SD") if "Assistenz" not in row[p]]
                for c in cands[:3]:
                    old = row[c].replace(" + Auffüllen", "")
                    suggestions.append({
                        "day": di, "msg": msg,
                        "description": f"{DAYS[di]}: {c}  {old} → SD: Assistenz Ulf",
                        "overrides": [{"person": c, "di": di, "value": "SD: Assistenz Ulf"}],
                    })

            # ── Lisa-OP-Assistenz fehlt ──────────────────────────────────────
            elif "Lisa OP → FD" in msg:
                cands = [p for p in tfa_in_shift("FD") if "Assistenz" not in row[p]]
                for c in cands[:3]:
                    old = row[c].replace(" + Auffüllen", "")
                    suggestions.append({
                        "day": di, "msg": msg,
                        "description": f"{DAYS[di]}: {c}  {old} → FD: Assistenz Lisa OP",
                        "overrides": [{"person": c, "di": di, "value": "FD: Assistenz Lisa OP"}],
                    })

            elif "Lisa OP → SD" in msg:
                cands = [p for p in tfa_in_shift("SD") if "Assistenz" not in row[p]]
                for c in cands[:3]:
                    old = row[c].replace(" + Auffüllen", "")
                    suggestions.append({
                        "day": di, "msg": msg,
                        "description": f"{DAYS[di]}: {c}  {old} → SD: Assistenz Lisa OP",
                        "overrides": [{"person": c, "di": di, "value": "SD: Assistenz Lisa OP"}],
                    })

            # ── Wilke-OP-Assistenz fehlt ─────────────────────────────────────
            elif "Wilke FD OP" in msg:
                cands = [p for p in TFAS if not _is_free(row[p]) and "Assistenz" not in row[p]]
                for c in cands[:3]:
                    old = row[c].replace(" + Auffüllen", "")
                    suggestions.append({
                        "day": di, "msg": msg,
                        "description": f"{DAYS[di]}: {c}  {old} → Assistenz Wilke OP",
                        "overrides": [{"person": c, "di": di, "value": "Assistenz Wilke OP"}],
                    })

    return suggestions


# ---------------------------------------------------------------------------
# Regel-Validierung
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Überstunden-Abbau-Rotation
# ---------------------------------------------------------------------------

def _load_ot_rotation() -> dict:
    if OT_ROTATION_PATH.exists():
        with open(OT_ROTATION_PATH, encoding="utf-8") as f:
            return json.load(f)
    return {"counts": {p: 0 for p in TFAS}}


def _save_ot_rotation(data: dict):
    content = json.dumps(data, indent=2, ensure_ascii=False)
    OT_ROTATION_PATH.write_text(content, encoding="utf-8")
    _github_push("overtime_rotation.json", content, "Auto-save: Überstunden-Rotation")


def _check_overtime(plan: dict, kw: int) -> list:
    """
    Prüft Mo–Fr: Ist ein Arzt wegen Urlaub/Krank abwesend UND sind ≥4 TFAs aktiv?
    Testet per echter Regelvalidierung ob jemand freigestellt werden kann.

    Ablauf pro Kandidat (nach fairem Rotationszähler sortiert):
      1. Kandidat → "Frei" simulieren → validieren → kein Fehler? → Vorschlag!
      2. Nur Anmeldungs-Lücken entstehen? → Ersatz im selben Shift suchen
         (Behandlungs-Person wechselt zu Anmeldung) → nochmals validieren → Vorschlag
         mit extra "swap"-Feld (Person A frei, Person B übernimmt Anmeldung)
      3. Andere Regeln (Assistenz Ulf/Lisa/Wilke) verletzt → Kandidat überspringen
    """
    from scheduler import week_dates
    from config import FEIERTAGE_HH_2026, ARZT_PATTERNS

    dates    = week_dates(kw)
    rotation = _load_ot_rotation()
    counts   = rotation.get("counts", {p: 0 for p in TFAS})

    ABSENT_STATES = {"Urlaub", "Krank"}

    def doctor_absent(val: str) -> bool:
        return val in ABSENT_STATES or val.startswith("Frei")

    def tfa_active(val: str) -> bool:
        skip = {"–", "Urlaub", "Feiertag", "SCHULE", "FREI_SCHUTZ",
                "Abschlussprüfung", "Krank", "ECVO Congress",
                "Notdienst (20–8 Uhr)", "Frei (ÜS-Abbau)"}
        return val not in skip and not val.startswith("Frei")

    def sim_plan(overrides: list[tuple]) -> dict:
        """Erstellt Plan-Kopie mit angewendeten (person, di, value)-Overrides."""
        p = {person: list(plan[person]) for person in ALL_PERSONS}
        for person, day_i, val in overrides:
            p[person][day_i] = val
        return p

    def day_issues(sp: dict) -> list[str]:
        return _validate_plan(sp, kw)["day"].get(str(di), [])

    suggestions = []

    for di in range(5):
        if dates[di] in FEIERTAGE_HH_2026:
            continue

        row = {p: _strip_note(str(plan[p][di])) for p in ALL_PERSONS}

        # Ärzte, die an diesem Tag normalerweise arbeiten, aber absent sind
        absent_docs = [
            a for a in ARZTE
            if doctor_absent(row[a]) and ARZT_PATTERNS[a][di] != "–"
        ]
        if not absent_docs:
            continue

        active_tfas = [p for p in TFAS if tfa_active(row[p])]
        if len(active_tfas) < 4:
            continue  # zu wenig Personal – kein Spielraum

        # Kandidaten: aktive TFAs, sortiert nach Rotationszähler (fairste Verteilung)
        pool = sorted(active_tfas, key=lambda p: (counts.get(p, 0), TFAS.index(p)))

        found = False
        for candidate in pool:
            cand_val = row[candidate]
            cand_shift = ("FD" if cand_val.startswith("FD")
                          else "SD" if cand_val.startswith("SD") else None)

            # ── Versuch 1: Kandidat direkt freistellen ─────────────────────
            issues = day_issues(sim_plan([(candidate, di, "Frei (ÜS-Abbau)")]))
            if not issues:
                suggestions.append({
                    "di": di, "day": DAYS[di],
                    "date": dates[di].strftime("%-d.%-m."),
                    "absent_doctors": absent_docs,
                    "active_tfa_count": len(active_tfas),
                    "suggested_person": candidate,
                    "suggested_count": counts.get(candidate, 0),
                    "current_shift": cand_val,
                    "swap": None,
                })
                found = True
                break

            # ── Versuch 2: Nur Anmeldungs-Lücken? → Ersatz suchen ──────────
            if not all("Anmeldung fehlt" in i for i in issues):
                continue  # Assistenz/Wilke/Lisa-Regeln betroffen → überspringen

            if not cand_shift:
                continue

            # Suche im selben Shift jemanden mit Behandlung (kein Assistenz)
            # als Anmeldungs-Ersatz; sortiert nach Rotationszähler für Fairness
            rep_pool = sorted(
                [p for p in active_tfas
                 if p != candidate
                 and row[p].startswith(cand_shift)
                 and "Behandlung" in row[p]
                 and "Assistenz" not in row[p]],
                key=lambda p: counts.get(p, 0)
            )
            for rep in rep_pool:
                new_rep_val = f"{cand_shift}: Anmeldung"
                issues2 = day_issues(sim_plan([
                    (candidate, di, "Frei (ÜS-Abbau)"),
                    (rep,       di, new_rep_val),
                ]))
                if not issues2:
                    suggestions.append({
                        "di": di, "day": DAYS[di],
                        "date": dates[di].strftime("%-d.%-m."),
                        "absent_doctors": absent_docs,
                        "active_tfa_count": len(active_tfas),
                        "suggested_person": candidate,
                        "suggested_count": counts.get(candidate, 0),
                        "current_shift": cand_val,
                        "swap": {
                            "person": rep,
                            "from":   row[rep],
                            "to":     new_rep_val,
                        },
                    })
                    found = True
                    break
            if found:
                break

    return suggestions


def _strip_note(val) -> str:
    """Entfernt optionalen Hinweis-Text (nach ' || ') aus dem Zellwert."""
    s = str(val) if val else "–"
    return s.split(" || ")[0] if " || " in s else s


def _validate_plan(plan: dict, kw: int) -> dict:
    from scheduler import week_dates
    from config    import FEIERTAGE_HH_2026

    dates  = week_dates(kw)
    issues = {"day": {}, "cell": {p: {} for p in ALL_PERSONS}}

    def v(val):    return _strip_note(val)  # Note-Teil ignorieren bei Validierung
    def shift(val):
        s = v(val)
        return "FD" if s.startswith("FD") else ("SD" if s.startswith("SD") else None)
    def working(val):
        s = v(val)
        return s not in ("–", "Urlaub", "Feiertag", "SCHULE", "FREI_SCHUTZ",
                         "Abschlussprüfung") and "Frei" not in s
    def add_day(di, msg):
        issues["day"].setdefault(str(di), []).append(msg)
    def add_cell(p, di, msg):
        issues["cell"][p].setdefault(str(di), []).append(msg)

    for di in range(5):
        if dates[di] in FEIERTAGE_HH_2026:
            continue
        row          = {p: v(plan[p][di]) for p in ALL_PERSONS}
        active_tfas  = [p for p in TFAS if working(row[p])]

        # Anmeldung
        fd_anm = [p for p in TFAS if "FD" in row[p] and "Anmeldung" in row[p]]
        sd_anm = [p for p in TFAS if "SD" in row[p] and "Anmeldung" in row[p]]
        # Warnung sobald ≥1 TFA in einem Shift aktiv ist — unabhängig von der Rolle.
        # (Frühere ">= 2 Behandlung"-Schwelle hat echte Lücken übersehen.)
        fd_active = [p for p in TFAS if "FD" in row[p] and working(row[p])]
        sd_active = [p for p in TFAS if "SD" in row[p] and working(row[p])]
        if fd_active and not fd_anm: add_day(di, "FD: Anmeldung fehlt")
        if sd_active and not sd_anm: add_day(di, "SD: Anmeldung fehlt")
        for p in fd_anm[1:]: add_cell(p, di, "FD: Anmeldung doppelt besetzt")
        for p in sd_anm[1:]: add_cell(p, di, "SD: Anmeldung doppelt besetzt")
        if len(fd_anm) > 1: add_day(di, f"FD: Anmeldung mehrfach ({', '.join(fd_anm)})")
        if len(sd_anm) > 1: add_day(di, f"SD: Anmeldung mehrfach ({', '.join(sd_anm)})")

        # Ulf-Assistenz
        ulf_present = row.get("Ulf", "–") not in ("–", "Urlaub", "Feiertag")
        fd_ulf = [p for p in TFAS if "FD" in row[p] and "Assistenz Ulf" in row[p]]
        sd_ulf = [p for p in TFAS if "SD" in row[p] and "Assistenz Ulf" in row[p]]
        if ulf_present:
            if not fd_ulf: add_day(di, "Ulf anwesend → FD: Assistenz fehlt")
            if not sd_ulf: add_day(di, "Ulf anwesend → SD: Assistenz fehlt")
            for p in fd_ulf[1:]: add_cell(p, di, "FD: Assistenz Ulf doppelt")
            for p in sd_ulf[1:]: add_cell(p, di, "SD: Assistenz Ulf doppelt")
        else:
            for p in fd_ulf + sd_ulf: add_cell(p, di, "Ulf Assistenz obwohl Ulf abwesend")

        # Lisa OP
        if "OP Ganztag" in row.get("Lisa", "–"):
            fd_l = [p for p in TFAS if "FD" in row[p] and "Assistenz Lisa OP" in row[p]]
            sd_l = [p for p in TFAS if "SD" in row[p] and "Assistenz Lisa OP" in row[p]]
            if not fd_l: add_day(di, "Lisa OP → FD: Assistenz fehlt")
            if not sd_l: add_day(di, "Lisa OP → SD: Assistenz fehlt")
            for p in fd_l[1:]: add_cell(p, di, "FD: Assistenz Lisa OP doppelt")
            for p in sd_l[1:]: add_cell(p, di, "SD: Assistenz Lisa OP doppelt")

        # Wilke OP
        if "FD OP" in row.get("Wilke", "–"):
            fw = [p for p in TFAS if "Assistenz Wilke OP" in row[p]]
            # Fr in ungeraden KWs: Imke übernimmt keinen Wilke-OP-Slot (Struktur-Ausnahme)
            wilke_fr_ungerade_ausnahme = (
                di == 4 and kw % 2 != 0
                and working(row.get("Imke", "–"))
            )
            if not fw and not wilke_fr_ungerade_ausnahme:
                add_day(di, "Wilke FD OP → Assistenz fehlt")
            for p in fw[1:]: add_cell(p, di, "Assistenz Wilke OP doppelt")

        # FD/SD-Balance
        fd_n = sum(1 for p in TFAS if shift(row[p]) == "FD")
        sd_n = sum(1 for p in TFAS if shift(row[p]) == "SD")
        if active_tfas and abs(fd_n - sd_n) >= 4:
            add_day(di, f"FD/SD-Ungleichgewicht: FD={fd_n}, SD={sd_n}")

    return issues


# ---------------------------------------------------------------------------
# Überstunden-Abbau-Routen
# ---------------------------------------------------------------------------

@app.route("/api/overtime/<int:kw>")
def api_overtime(kw):
    """Gibt Überstunden-Abbau-Vorschläge für eine KW zurück (kein Auth nötig)."""
    if not _plan_cache:
        _generate_all()
    overrides   = _load_overrides()
    plan        = _plan_for_kw(kw, overrides)
    suggestions = _check_overtime(plan, kw)
    return jsonify(suggestions)


@app.route("/api/overtime/apply", methods=["POST"])
@require_auth
def api_overtime_apply():
    """
    Trägt einen Überstunden-Freitag ein und erhöht den Rotationszähler der Person.
    Body: { kw, di, person }
    """
    if not _plan_cache:
        _generate_all()
    data   = request.get_json()
    kw     = int(data["kw"])
    di     = int(data["di"])
    person = data["person"]

    if person not in ALL_PERSONS:
        return jsonify({"ok": False, "error": "Unbekannte Person"}), 400

    swap = data.get("swap")  # optional: {"person": ..., "value": ...}

    # Freier Tag als Override eintragen
    overrides = _load_overrides()
    overrides.setdefault(str(kw), {}).setdefault(person, {})[str(di)] = "Frei (ÜS-Abbau)"
    # Optional: Anmeldungs-Ersatz übernehmen
    if swap and swap.get("person") and swap.get("value"):
        overrides.setdefault(str(kw), {}).setdefault(swap["person"], {})[str(di)] = swap["value"]
    _save_overrides(overrides)

    # Rotationszähler der Person erhöhen
    rotation = _load_ot_rotation()
    rotation.setdefault("counts", {})[person] = rotation["counts"].get(person, 0) + 1
    _save_ot_rotation(rotation)

    # Aktualisierte Vorschläge + Validierung zurückgeben
    plan        = _plan_for_kw(kw, overrides)
    issues      = _validate_plan(plan, kw)
    suggestions = _check_overtime(plan, kw)

    return jsonify({
        "ok":           True,
        "issues":       issues,
        "ot_suggestions": suggestions,
        "total_issues": sum(len(v) for v in issues["day"].values()),
    })


# ---------------------------------------------------------------------------
# Start
# ---------------------------------------------------------------------------

def _get_local_ip() -> str:
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "unbekannt"


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5050))
    print("Synchronisiere Overrides von GitHub …")
    _sync_from_github()
    print("Generiere Plan KW 17–52 …")
    _generate_all()
    ip = _get_local_ip()
    print(f"""
╔══════════════════════════════════════════════════════╗
  Dienstplan-App gestartet

  Auf diesem Computer:   http://localhost:{port}
  Im Praxis-Netzwerk:    http://{ip}:{port}

  Mitarbeiter können den Plan NUR ansehen.
  Bearbeitungs-PIN: {EDIT_PIN}
  (änderbar mit: export EDIT_PIN=neupin)
╚══════════════════════════════════════════════════════╝
""")
    app.run(host="0.0.0.0", port=port, debug=False, use_reloader=False)
