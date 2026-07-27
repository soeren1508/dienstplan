"""
vacations.py — Liest Urlaubsplanung 2026.xlsx.

Rückgabe von load_vacations():
    { vorname: set[date] }
    z.B. { "Kristin": {date(2026, 6, 22), ...}, ... }
"""

from datetime import date, timedelta
from pathlib import Path
from typing import Dict, Set

import openpyxl

from config import NACHNAME_ZU_VORNAME

# Urlaubsplanung: Col 4 = 01.01.2026 (Donnerstag)
_START_COL = 4
_START_DATE = date(2026, 1, 1)
# Personen stehen in Zeilen 7–18, Spalte A
_PERSON_ROWS = range(7, 19)


def _load_marker(path: str | Path, marker: str) -> Dict[str, Set[date]]:
    """Parst die Urlaubsplanung und gibt Tage mit dem angegebenen Kürzel
    (z.B. 'U' für Urlaub, 'K' für Krank) pro Person zurück."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    max_col = ws.max_column

    result: Dict[str, Set[date]] = {}

    for ri in _PERSON_ROWS:
        row = next(ws.iter_rows(
            min_row=ri, max_row=ri,
            min_col=1, max_col=max_col,
            values_only=True
        ))
        nachname = row[0]
        if not nachname:
            continue
        nachname = str(nachname).strip()
        vorname = NACHNAME_ZU_VORNAME.get(nachname)
        if not vorname:
            continue

        days: Set[date] = set()
        for ci, val in enumerate(row[_START_COL - 1:], _START_COL):
            if isinstance(val, str) and val.strip().upper() == marker:
                d = _START_DATE + timedelta(days=ci - _START_COL)
                days.add(d)

        result[vorname] = days

    return result


def load_vacations(path: str | Path) -> Dict[str, Set[date]]:
    """Parst die Urlaubsplanung, gibt Urlaubstage ('U'-Einträge) pro Person zurück."""
    return _load_marker(path, "U")


def load_krank(path: str | Path) -> Dict[str, Set[date]]:
    """Parst die Urlaubsplanung, gibt geplante Krank-Tage ('K'-Einträge) pro Person zurück."""
    return _load_marker(path, "K")


def vacations_for_week(
    all_vacations: Dict[str, Set[date]],
    kw_dates: list[date]
) -> Dict[str, list[bool]]:
    """
    Gibt für jede Person eine Liste [mo, di, mi, do, fr, sa] zurück,
    True = Urlaub an diesem Tag.
    kw_dates: Liste von 6 date-Objekten (Mo–Sa der Woche).
    """
    result: Dict[str, list[bool]] = {}
    for person, days in all_vacations.items():
        result[person] = [d in days for d in kw_dates]
    return result
